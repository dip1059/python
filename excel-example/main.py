import openpyxl as xl
import openpyxl.styles.fills as fills
from openpyxl.styles import PatternFill, Font, Color

STATUS_COL = 5
ACTIVITY_COL = 7


def change_cell_value(sheet, row, col, value):
    cell = sheet.cell(row, col)
    cell.value = value
    cell.font = Font(color="000000")
    cell.fill = PatternFill(fills.FILL_PATTERN_DARKUP, fgColor="000066CC")


def process_data(sheet):
    rows = sheet.max_row
    for row in range(2, rows+1):
        status = sheet.cell(row, STATUS_COL).value
        if status == 1:
            activity = "Active"
        elif status == 2:
            activity = "Suspended"
        else:
            activity = "Inactive"
        change_cell_value(sheet, row, ACTIVITY_COL, activity)


def main():
    file = xl.load_workbook('export_result.xlsx')
    sheet = file['export_result']
    change_cell_value(sheet, 1, ACTIVITY_COL, "activity")
    process_data(sheet)
    file.save("out.xlsx")


main()
